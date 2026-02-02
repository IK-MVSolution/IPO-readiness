from __future__ import annotations

import os
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import xlrd
from openpyxl import Workbook, load_workbook

YEARS = [2563, 2564, 2565, 2566, 2567]

STANDARD_LAYOUT = {
    "ratio": {
        "columns": ["B", "C", "D", "E", "F", "G"],
        "rows": {
            "roa": 7,
            "roe": 8,
            "current_ratio": 12,
            "debt_to_equity": 22,
            "debt_to_assets": 23,
            "gross_profit_margin_pct": 9,  # rarely present but captured if exists
        },
    },
    "income": {
        "columns": ["B", "D", "F", "H", "J"],
        "rows": {
            "total_revenue": 7,
            "gross_profit": 9,
            "net_profit": 14,
        },
    },
    "balance": {
        "columns": ["B", "D", "F", "H", "J"],
        "rows": {
            "total_assets": 11,
            "total_liabilities": 14,
            "shareholders_equity": 15,
        },
    },
}

LEGACY_LAYOUT = {
    "ratio": {
        "columns": ["C", "D", "E", "F", "G"],
        "rows": {
            "roa": 6,
            "roe": 7,
            "gross_profit_margin_pct": 8,
            "net_profit_margin_pct": 9,
            "current_ratio": 11,
            "debt_to_assets": 16,
            "debt_to_equity": 18,
        },
    },
    "income": {
        "columns": ["C", "E", "G", "I", "K"],
        "rows": {
            "total_revenue": 6,
            "net_profit": 12,
        },
    },
    "balance": {
        "columns": ["C", "E", "G", "I", "K"],
        "rows": {
            "total_assets": 9,
            "shareholders_equity": 13,
        },
    },
}

LAYOUTS = {
    "standard": STANDARD_LAYOUT,
    "legacy": LEGACY_LAYOUT,
}


def parse_financial_files(workbooks):
    files = workbooks if isinstance(workbooks, list) else [workbooks]
    aggregated = {
        "years": YEARS,
        "balance_sheet": {},
        "income_statement": {},
        "ratios": {},
        "company_name": None,
    }
    
    print("\n" + "="*80)
    print("📊 เริ่มการประมวลผลไฟล์ทางการเงิน")
    print("="*80)
    
    for idx, file in enumerate(files):
        filename = getattr(file, "filename", "") or getattr(file, "name", f"file_{idx+1}")
        print(f"\n📁 ไฟล์ที่ {idx+1}: {filename}")
        
        workbook, layout_key = _load_workbook_from_upload(file)
        print(f"   Layout: {layout_key}")
        print(f"   Sheets: {[sheet.title for sheet in workbook.worksheets]}")
        
        # Extract company name from first file if not already set
        if aggregated["company_name"] is None:
            aggregated["company_name"] = _extract_company_name(workbook)
            if aggregated["company_name"]:
                print(f"   🏢 ชื่อบริษัท: {aggregated['company_name']}")
            else:
                print(f"   ⚠️  ไม่พบชื่อบริษัท")
        
        extracted = _extract_from_workbook(workbook, layout_key)
        _merge_sections(aggregated["balance_sheet"], extracted["balance_sheet"])
        _merge_sections(aggregated["income_statement"], extracted["income_statement"])
        _merge_sections(aggregated["ratios"], extracted["ratios"])
    
    # Summary
    print("\n" + "-"*80)
    print("📝 สรุปข้อมูลที่ดึงได้:")
    print("-"*80)
    
    for section_name, section_data in [
        ("งบดุล (Balance Sheet)", aggregated["balance_sheet"]),
        ("งบกำไรขาดทุน (Income Statement)", aggregated["income_statement"]),
        ("อัตราส่วน (Ratios)", aggregated["ratios"])
    ]:
        print(f"\n{section_name}:")
        if section_data:
            for metric_name, values in section_data.items():
                years_found = sorted(values.keys())
                print(f"  ✓ {metric_name}: {len(years_found)} ปี {years_found}")
        else:
            print(f"  ⚠️  ไม่มีข้อมูล")
    
    print("\n" + "="*80)
    print("✅ ประมวลผลเสร็จสิ้น")
    print("="*80 + "\n")
    
    return aggregated


def _extract_company_name(workbook) -> Optional[str]:
    """
    Extract company name from Excel workbook.
    Typically found in row 1, columns A-G of the first or ratio sheet.
    """
    # Try ratio sheet first, then first sheet
    sheets_to_check = []
    for sheet in workbook.worksheets:
        sheet_type = _detect_sheet_type(sheet)
        if sheet_type == "ratio":
            sheets_to_check.insert(0, sheet)  # Prioritize ratio sheet
        elif not sheets_to_check:
            sheets_to_check.append(sheet)  # Fallback to first sheet
    
    if not sheets_to_check:
        return None
    
    # Check first few rows for company name, prioritizing row 1
    for sheet in sheets_to_check[:2]:  # Check max 2 sheets
        for row_num in [1, 2, 3]:  # Check rows 1-3, prioritize 1
            for col_letter in ["C", "B", "A", "D", "E", "F", "G"]:  # Prioritize C, B, A
                cell_value = sheet[f"{col_letter}{row_num}"].value
                if cell_value and isinstance(cell_value, str):
                    # Clean up the value
                    cleaned = str(cell_value).strip()
                    # Check if it looks like a company name
                    if len(cleaned) > 5 and (
                        any(ord(char) >= 0x0E00 and ord(char) <= 0x0E7F for char in cleaned) or  # Contains Thai
                        "จำกัด" in cleaned or 
                        "บริษัท" in cleaned or
                        "มหาชน" in cleaned or
                        "Company" in cleaned.title() or
                        "Ltd" in cleaned or
                        "Public" in cleaned.title()
                    ):
                        # Extract just the company name part
                        # Remove common prefixes/headers
                        result = cleaned
                        if "บริษัท" in result:
                            # Extract from "บริษัท" onwards
                            idx = result.find("บริษัท")
                            result = result[idx:]
                        elif "Company" in result.title():
                            idx = result.lower().find("company")
                            result = result[idx:]
                        
                        # Remove header text before company name
                        for prefix in ["อัตราส่วนทางการเงิน - ", "อัตราส่วนทางการเงินที่สำคัญ ", "อัตราส่วน - "]:
                            if prefix in result:
                                result = result.replace(prefix, "").strip()
                        
                        print(f"🏢 DEBUG: Found company name in {col_letter}{row_num}: {result}")
                        return result
    return None



def _extract_from_workbook(workbook, layout_key: str) -> Dict[str, Dict[str, Dict[int, float]]]:
    layout = LAYOUTS.get(layout_key, STANDARD_LAYOUT)
    sections = {
        "balance_sheet": {},
        "income_statement": {},
        "ratios": {},
    }
    for sheet in workbook.worksheets:
        sheet_type = _detect_sheet_type(sheet)
        if not sheet_type:
            continue
        config = layout.get(sheet_type)
        if not config:
            continue
        extracted = _extract_by_cells(sheet, config["rows"], config["columns"])
        if sheet_type == "ratio":
            sections["ratios"] = extracted
        elif sheet_type == "income":
            sections["income_statement"] = extracted
        elif sheet_type == "balance":
            sections["balance_sheet"] = extracted
    _fill_derived_values(sections)
    return sections


def _merge_sections(target: Dict[str, Dict[int, float]], source: Dict[str, Dict[int, float]]) -> None:
    for key, series in source.items():
        existing = target.setdefault(key, {})
        existing.update(series)


def _detect_sheet_type(sheet) -> Optional[str]:
    keywords = []
    for row in sheet.iter_rows(min_row=1, max_row=8, max_col=4):
        for cell in row:
            if cell.value:
                keywords.append(str(cell.value).lower())
    text = " ".join(keywords)
    if "อัตราส่วน" in text or "ratio" in text:
        return "ratio"
    if "กำไร" in text or "income" in text or "profit" in text:
        return "income"
    if "ฐานะ" in text or "balance" in text or "สินทรัพย์รวม" in text:
        return "balance"
    # fallback based on sheet name
    name = sheet.title.lower()
    if any(word in name for word in ["ratio", "อัตราส่วน"]):
        return "ratio"
    if any(word in name for word in ["income", "profit", "กำไร"]):
        return "income"
    if any(word in name for word in ["balance", "สินทรัพย์", "ฐานะ"]):
        return "balance"
    return None


def _extract_by_cells(sheet, rows_map: Dict[str, int], columns: List[str]) -> Dict[str, Dict[int, float]]:
    data: Dict[str, Dict[int, float]] = {}
    for key, row in rows_map.items():
        data[key] = {}
        for col, year in zip(columns, YEARS):
            cell_value = sheet[f"{col}{row}"].value
            value = _coerce_number(cell_value)
            if value is not None:
                data[key][year] = value
    return data


def _coerce_number(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(",", "")
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def _fill_derived_values(sections: Dict[str, Dict[str, Dict[int, float]]]) -> None:
    _derive_gross_profit(sections)
    _derive_total_liabilities(sections)


def _derive_gross_profit(sections: Dict[str, Dict[str, Dict[int, float]]]) -> None:
    income = sections["income_statement"]
    if income.get("gross_profit"):
        return
    revenue = income.get("total_revenue", {})
    ratios = sections["ratios"].get("gross_profit_margin_pct", {})
    derived: Dict[int, float] = {}
    for year, rev in revenue.items():
        margin = ratios.get(year)
        if margin is None:
            continue
        derived[year] = rev * (margin / 100.0)
    if derived:
        income["gross_profit"] = derived


def _derive_total_liabilities(sections: Dict[str, Dict[str, Dict[int, float]]]) -> None:
    balance = sections["balance_sheet"]
    if balance.get("total_liabilities"):
        return
    assets = balance.get("total_assets", {})
    equity = balance.get("shareholders_equity", {})
    derived: Dict[int, float] = {}
    for year, asset in assets.items():
        eq = equity.get(year)
        if eq is None:
            continue
        derived_value = asset - eq
        derived[year] = derived_value
    if derived:
        balance["total_liabilities"] = derived


def _load_workbook_from_upload(file_obj) -> Tuple[Workbook, str]:
    filename = getattr(file_obj, "filename", "") or getattr(file_obj, "name", "")
    extension = os.path.splitext(filename.lower())[1]
    raw_bytes = _read_upload_bytes(file_obj)
    if extension == ".xls":
        return _convert_xls_to_openpyxl(raw_bytes), "legacy"
    return load_workbook(BytesIO(raw_bytes), data_only=True), "standard"


def _read_upload_bytes(file_obj) -> bytes:
    stream = getattr(file_obj, "stream", file_obj)
    if hasattr(stream, "seek"):
        stream.seek(0)
    data = stream.read()
    if hasattr(stream, "seek"):
        stream.seek(0)
    return data


def _convert_xls_to_openpyxl(raw_bytes: bytes) -> Workbook:
    book = xlrd.open_workbook(file_contents=raw_bytes)
    workbook = Workbook()
    active = workbook.active
    if book.nsheets == 0:
        active.title = "Sheet1"
        return workbook

    first_sheet = book.sheet_by_index(0)
    active.title = _safe_sheet_title(first_sheet.name)
    _copy_xlrd_sheet(first_sheet, active)

    for index in range(1, book.nsheets):
        source = book.sheet_by_index(index)
        target = workbook.create_sheet(title=_safe_sheet_title(source.name))
        _copy_xlrd_sheet(source, target)
    return workbook


def _copy_xlrd_sheet(source_sheet, target_sheet) -> None:
    for row in range(source_sheet.nrows):
        for col in range(source_sheet.ncols):
            target_sheet.cell(row=row + 1, column=col + 1, value=source_sheet.cell_value(row, col))


def _safe_sheet_title(title: str) -> str:
    base = title or "Sheet"
    cleaned = base.replace("/", "-")
    return cleaned[:31]
